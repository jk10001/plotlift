from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook

from app.exports import csv_bytes, export_fieldnames, export_rows, xlsx_bytes
from app.models import (
    Asset,
    CalibratedAxis,
    CalibrationPoint,
    CalibrationState,
    ChartValue,
    CropState,
    NormBBox,
    NormPoint,
    PixelBBox,
    PixelPoint,
    PromptMetadata,
    RunSettings,
    RunState,
    SeriesPoint,
    SeriesState,
)


def test_base_export_omits_debug_columns() -> None:
    rows = export_rows(_export_state(), include_debug=False)
    assert rows
    assert list(rows[0].keys()) == export_fieldnames(include_debug=False)
    assert "series_name" in rows[0]
    for field in _debug_only_fields():
        assert field not in rows[0]

    csv_rows = list(csv.DictReader(StringIO(csv_bytes(_export_state(), include_debug=False).decode("utf-8"))))
    assert csv_rows
    for field in _debug_only_fields():
        assert field not in csv_rows[0]


def test_debug_export_includes_internal_and_coordinate_columns() -> None:
    rows = export_rows(_export_state(), include_debug=True)
    assert rows
    assert list(rows[0].keys()) == export_fieldnames(include_debug=True)
    for field in _debug_only_fields():
        assert field in rows[0]
    assert rows[0]["run_id"] == "export-test"
    assert rows[0]["llm_series_name"] == "LLM Name"
    assert rows[0]["segment_index"] == 2
    assert rows[0]["point_index"] == 7
    assert rows[0]["x_axis_id"] == "x_flow"
    assert rows[0]["y_axis_id"] == "y_head"
    assert rows[0]["x_axis_unit"] == "L/s"
    assert rows[0]["y_axis_unit"] == "m"
    assert rows[0]["crop_image_norm_x"] == 400
    assert rows[0]["full_image_px_x"] == 120


def test_xlsx_data_sheet_uses_matching_export_columns() -> None:
    base_workbook = load_workbook(BytesIO(xlsx_bytes(_export_state(), include_debug=False)), read_only=True)
    debug_workbook = load_workbook(BytesIO(xlsx_bytes(_export_state(), include_debug=True)), read_only=True)

    assert base_workbook.sheetnames == ["data"]
    assert "metadata" in debug_workbook.sheetnames
    assert "series_summary" in debug_workbook.sheetnames
    assert "axes" in debug_workbook.sheetnames
    assert next(base_workbook["data"].iter_rows(values_only=True)) == tuple(export_fieldnames(include_debug=False))
    assert next(debug_workbook["data"].iter_rows(values_only=True)) == tuple(export_fieldnames(include_debug=True))


def _debug_only_fields() -> list[str]:
    return [
        "run_id",
        "llm_series_name",
        "segment_index",
        "point_index",
        "x_axis_id",
        "y_axis_id",
        "x_axis_unit",
        "y_axis_unit",
        "crop_image_norm_x",
        "crop_image_norm_y",
        "crop_image_px_x",
        "crop_image_px_y",
        "full_image_px_x",
        "full_image_px_y",
    ]


def _export_state() -> RunState:
    return RunState(
        run_id="export-test",
        stage="complete",
        settings=RunSettings(model_id="gpt-5.4-mini"),
        prompt_metadata=PromptMetadata(prompt_version="test", prompt_hash="abc"),
        upload_filename="chart.png",
        original_path="uploads/chart.png",
        canonical_image=Asset(path="canonical.png", width=300, height=200),
        crop=CropState(
            bbox_full_norm=NormBBox(left=0, top=0, right=999, bottom=999),
            bbox_full_px=PixelBBox(left=100, top=50, right=300, bottom=250),
            image=Asset(path="crop.png", width=200, height=200),
            approved=True,
        ),
        calibration=CalibrationState(
            calibrated_axes=[
                CalibratedAxis(
                    axis_id="x_flow",
                    direction="x",
                    name="Flow",
                    unit="L/s",
                    location_description="bottom axis",
                    approved=True,
                    points=[
                        CalibrationPoint(label="x1", crop_image_norm=NormPoint(x=0, y=900), chart_value=ChartValue(value_raw="0")),
                        CalibrationPoint(label="x2", crop_image_norm=NormPoint(x=999, y=900), chart_value=ChartValue(value_raw="10")),
                    ],
                ),
                CalibratedAxis(
                    axis_id="y_head",
                    direction="y",
                    name="Head",
                    unit="m",
                    location_description="left axis",
                    approved=True,
                    points=[
                        CalibrationPoint(label="y1", crop_image_norm=NormPoint(x=0, y=999), chart_value=ChartValue(value_raw="0")),
                        CalibrationPoint(label="y2", crop_image_norm=NormPoint(x=0, y=0), chart_value=ChartValue(value_raw="100")),
                    ],
                ),
            ]
        ),
        series=[
            SeriesState(
                id="series-1",
                name="Series A",
                llm_series_name="LLM Name",
                x_axis_id="x_flow",
                y_axis_id="y_head",
                points=[
                    SeriesPoint(
                        point_index=7,
                        segment_index=2,
                        crop_image_norm=NormPoint(x=400, y=600),
                        crop_image_px=PixelPoint(x=20, y=30),
                        chart_x=ChartValue(value_raw="1", unit="L/s"),
                        chart_y=ChartValue(value_raw="2", unit="m"),
                    )
                ],
            )
        ],
    )

from __future__ import annotations

import csv
import io
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from . import coordinates
from .models import ChartValue, RunState, SeriesPoint

BASE_EXPORT_COLUMNS = [
    "series_name",
    "x",
    "y",
    "x_value_type",
    "y_value_type",
    "x_unit",
    "y_unit",
    "x_axis_name",
    "y_axis_name",
]

DEBUG_EXPORT_COLUMNS = [
    "run_id",
    "llm_series_name",
    "segment_index",
    "point_index",
    "x_axis_id",
    "y_axis_id",
    "x_axis_unit",
    "y_axis_unit",
    "crop_image_norm_x",
    "crop_image_norm_y",
    "crop_image_px_x",
    "crop_image_px_y",
    "full_image_px_x",
    "full_image_px_y",
]


def chart_value_export(value: ChartValue | None) -> Any:
    if value is None:
        return None
    if value.value_type == "datetime":
        return value.parsed_datetime or value.value_raw
    if value.value_type == "number":
        return value.parsed_value
    return value.value_raw


def export_rows(state: RunState, include_debug: bool = False) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for series in state.series:
        x_axis = state.calibration.axis_by_id(series.x_axis_id)
        y_axis = state.calibration.axis_by_id(series.y_axis_id)
        for point in series.points:
            row: dict[str, Any] = {
                "series_name": series.name,
                "x": chart_value_export(point.chart_x),
                "y": chart_value_export(point.chart_y),
                "x_value_type": point.chart_x.value_type if point.chart_x else None,
                "y_value_type": point.chart_y.value_type if point.chart_y else None,
                "x_unit": point.chart_x.unit if point.chart_x else None,
                "y_unit": point.chart_y.unit if point.chart_y else None,
                "x_axis_name": x_axis.name if x_axis else None,
                "y_axis_name": y_axis.name if y_axis else None,
            }
            if include_debug:
                row.update(
                    {
                        "run_id": state.run_id,
                        "llm_series_name": series.llm_series_name,
                        "segment_index": point.segment_index,
                        "point_index": point.point_index,
                        "x_axis_id": series.x_axis_id,
                        "y_axis_id": series.y_axis_id,
                        "x_axis_unit": x_axis.unit if x_axis else None,
                        "y_axis_unit": y_axis.unit if y_axis else None,
                    }
                )
                row.update(_debug_columns(state, point))
            rows.append(row)
    return rows


def export_fieldnames(include_debug: bool = False) -> list[str]:
    return BASE_EXPORT_COLUMNS + (DEBUG_EXPORT_COLUMNS if include_debug else [])


def csv_bytes(state: RunState, include_debug: bool = False) -> bytes:
    rows = export_rows(state, include_debug=include_debug)
    output = io.StringIO()
    fieldnames = export_fieldnames(include_debug=include_debug)
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8")


def xlsx_bytes(state: RunState, include_debug: bool = False) -> bytes:
    rows = export_rows(state, include_debug=include_debug)
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    fieldnames = export_fieldnames(include_debug=include_debug)
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(field) for field in fieldnames])
    for index, field in enumerate(fieldnames, start=1):
        ws.column_dimensions[get_column_letter(index)].width = max(12, min(36, len(field) + 3))

    if not include_debug:
        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    meta = wb.create_sheet("metadata")
    metadata = {
        "run_id": state.run_id,
        "created_at": state.created_at,
        "updated_at": state.updated_at,
        "model_id": state.settings.model_id,
        "image_detail": state.settings.image_detail,
        "reasoning_effort": state.settings.reasoning_effort,
        "mock_mode": state.settings.mock_mode,
        "prompt_version": state.prompt_metadata.prompt_version,
        "prompt_hash": state.prompt_metadata.prompt_hash,
        "upload_filename": state.upload_filename,
    }
    for key, value in metadata.items():
        meta.append([key, value])

    summary = wb.create_sheet("series_summary")
    summary.append(["series_name", "llm_series_name", "x_axis_id", "y_axis_id", "points", "confidence", "visual_description"])
    for series in state.series:
        summary.append([series.name, series.llm_series_name, series.x_axis_id, series.y_axis_id, len(series.points), series.confidence, series.visual_description])

    axes = wb.create_sheet("axes")
    axes.append(["axis_id", "direction", "name", "unit", "quantity", "location_description", "approved"])
    for axis in state.calibration.calibrated_axes:
        axes.append([axis.axis_id, axis.direction, axis.name, axis.unit, axis.quantity, axis.location_description, axis.approved])

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def create_run_archive(run_root: Path, target_path: Path) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(target_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in run_root.rglob("*"):
            if path.is_file() and path.resolve() != target_path.resolve():
                archive.write(path, path.relative_to(run_root))


def _debug_columns(state: RunState, point: SeriesPoint) -> dict[str, Any]:
    debug: dict[str, Any] = {
        "crop_image_norm_x": point.crop_image_norm.x,
        "crop_image_norm_y": point.crop_image_norm.y,
        "crop_image_px_x": point.crop_image_px.x if point.crop_image_px else None,
        "crop_image_px_y": point.crop_image_px.y if point.crop_image_px else None,
    }
    if point.crop_image_px and state.crop:
        full = coordinates.crop_to_full_point(point.crop_image_px, state.crop.bbox_full_px)
        debug["full_image_px_x"] = full.x
        debug["full_image_px_y"] = full.y
    return debug
